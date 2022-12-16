"""
Author: Ronin Sharma
Date: December 14, 2022
"""

import pandas as pd
from liberal_studies_checker import LiberalStudiesChecker
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from constants import *

FILL_CELL_RED = PatternFill(patternType='solid', fgColor='C64747')

class RequirementChecker:

    def __init__(self, excel_data, course_mapping, excel_file_name):
        """
        Attributes:
            excel_data (ExcelData): object with the excel data functions 
            course_mapping (dict): stores the mapping between requirements and courses
            excel_file_name (string): checklist's excel file name
            multi_option_map (dict): requirements with multiple courses that
            ece_foundation (list): the possible ece foundation courses
            ece_foundation_taken (list): the ece foundation courses taken
            ls_classes (list): the liberal studies courses taken
            num_ls_credits (int): the number of credits from liberal studies courses
            num_aae_credits (int): the number of credits for advisor-approved elective courses
        """
        
        self.excel_data = excel_data
        self.course_mapping = course_mapping
        self.excel_file_name = excel_file_name
        self.multi_option_map = MULTI_OPTION
        self.ece_foundation = ECE_FOUNDATION
        self.ece_foundation_taken = []
        self.ls_classes = []
        self.num_ls_credits = 0
        self.num_aae_credits = 0
    
    def load_excel_file(self):
        """
        Loads the excel file and creates the new sheet.
        """
        
        self.excel_book = load_workbook(self.excel_file_name)
        self.excel_file = pd.ExcelFile(self.excel_file_name)
        # Check if there's an extra sheet to delete
        excel_sheet_names = self.excel_book.get_sheet_names()
        if (len(excel_sheet_names) > 1):
            excel_sheet = self.excel_book.get_sheet_by_name(excel_sheet_names[1])
            self.excel_book.remove_sheet(excel_sheet)

        # Duplicate input sheet
        excel_sheet = self.excel_book.get_sheet_by_name(excel_sheet_names[0])
        self.excel_book.copy_worksheet(excel_sheet)
        self.save_excel_file()

        # Reload to get the new file name
        self.excel_book = load_workbook(self.excel_file_name)
        self.excel_file = pd.ExcelFile(self.excel_file_name)
        new_sheet_name = self.excel_file.sheet_names[-1]
        self.excel_sheet = self.excel_book[new_sheet_name]

    def save_excel_file(self):
        """
        Saves the excel file
        """
        
        self.excel_book.save(self.excel_file_name)
    
    def update_cell_color(self, excel_data):
        """
        Marks the specified excel cell (default color is red).
        """
        
        for cell_letter in excel_data['letters']:
            self.excel_sheet[f"{cell_letter}{excel_data['num']}"].fill = FILL_CELL_RED
    
    def basic_verification(self, course_data, excel_data):
        """
        Verifies passing grade and completed checklist row.
        """
        
        if (course_data.get_grade() in ['U', 'F', 'D-', 'D', 'D+'] or course_data.get_missing()):
            valid_course = False
        else:
            valid_course = True
        
        if (not valid_course):
            self.update_cell_color(excel_data)

    def check_multi_option_course_requirement(self, course_data, excel_data):
        """
        Verifies that at least one course out of all options was taken.
        """
        
        valid_course = (course_data.get_name() in self.multi_option_map[course_data.get_requirement()])
        
        if (not valid_course):
            self.update_cell_color(excel_data)
    
    def check_3000_4000_ece(self, course_data, excel_data):
        """
        Verifies 3000+ and 4000+ ECE course requirement.
        """
        
        valid_course = False

        if (not course_data.get_missing()):
            course_name = course_data.get_name()
            course_level = int(course_name.split(' ')[1][0])
            course_requirement_level = int(course_data.get_requirement()[0])
            valid_course = (course_level >= course_requirement_level and course_name not in INVALID_UPPER_LEVEL_ECE)

        if (not valid_course):
            self.update_cell_color(excel_data)

    def check_ece_foundation(self):
        """
        Verifies ECE foundation courses.
        """
        
        # Check total courses taken
        num_taken = len(self.ece_foundation_taken)
        
        if (num_taken < 3):
            print('YOU NEED TO TAKE THREE CORE ECE 3000 COURSES')
        
        if ('ECE 3030' not in self.ece_foundation_taken and 'ECE 3150' not in self.ece_foundation_taken):
            print('YOU NEED TO TAKE EITHER ECE 3030 OR ECE 3150')
        
        if ('ECE 3100' not in self.ece_foundation_taken and 'ECE 3250' not in self.ece_foundation_taken):
            print('YOU NEED TO TAKE EITHER ECE 3100 OR ECE 3250')

    def check_requirements(self):
        """
        Checks all requirements and updates the excel file based on failed requirements.
        """
        
        for key in self.course_mapping:
            for course in self.course_mapping[key]:
                course_name = course.get_name()
                course_requirement = course.get_requirement()
                excel_data = self.excel_data.get_excel_data_by_course_id(course.get_id())
                if (course_requirement in ['FWS', 'AAE', 'OTE', 'ENGRD'] or course_name == 'PE'):
                    self.basic_verification(course, excel_data)
                elif (course_requirement == 'LS'):
                    self.num_ls_credits += course.get_credits()
                    self.ls_classes.append(course_name)
                elif (course_requirement in ['3000+', '4000+']):
                    self.basic_verification(course, excel_data)
                    self.check_3000_4000_ece(course, excel_data)
                elif (course_requirement in ['CDE', 'ENGRI']):
                    self.basic_verification(course, excel_data)
                    self.check_multi_option_course_requirement(course, excel_data)
                elif (course_name in self.ece_foundation and not (course.get_grade() in ['U', 'F', 'D-', 'D', 'D+'] or course.get_missing())):
                    self.ece_foundation_taken.append(course_name)
                elif (course_requirement == 'AAE'):
                    self.num_aae_credits += course.get_credits()
                elif course.check_valid_name():
                    self.basic_verification(course, excel_data)
        
        self.check_ece_foundation()

        if (self.num_aae_credits < 6):
            print('AT LEAST 6 CREDITS OF AAE COURSES ARE REQUIRED')

        if (self.num_ls_credits < 18):
            print('AT LEAST 18 CREDITS OF LS COURSES ARE REQUIRED')

        ls_checker = LiberalStudiesChecker(self.ls_classes)
        ls_checker.scrape_data()
        ls_result = ls_checker.check_ls_data()
        print(ls_result)

    def run(self):
        """
        Loads the excel file and checks all requirements.
        """
        
        self.load_excel_file()
        self.check_requirements()
